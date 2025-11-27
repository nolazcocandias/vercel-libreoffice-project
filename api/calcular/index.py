
import json
import random
import openpyxl
import subprocess
import os
from http.server import BaseHTTPRequestHandler

class handler(BaseHTTPRequestHandler):
    # --- CORS headers (siempre) ---
    def _set_cors_headers(self):
        # Para probar: permitir cualquier origen. Luego puedes cambiar a tu dominio específico:
        # 'https://nolazcocandias.github.io'
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        # Incluye los más comunes para fetch y navegadores modernos
        self.send_header('Access-Control-Allow-Headers', 'Content-Type, Accept, Origin, X-Requested-With')
        self.send_header('Access-Control-Max-Age', '86400')  # cache preflight 24h

    def _json_response(self, status_code, payload):
        self.send_response(status_code)
        self._set_cors_headers()
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.end_headers()
        self.wfile.write(json.dumps(payload).encode('utf-8'))

    # --- Preflight ---
    def do_OPTIONS(self):
        # Responder SIEMPRE con 200 y los headers CORS
        self.send_response(200)
        self._set_cors_headers()
        # Si quisieras limitar métodos/headers específicos, ya están en _set_cors_headers()
        self.end_headers()

    # (Opcional) algunos clientes prueban HEAD
    def do_HEAD(self):
        self.send_response(200)
        self._set_cors_headers()
        self.end_headers()

    # --- Salud (GET) ---
    def do_GET(self):
        # Devuelve un ok simple
        self._json_response(200, {"status": "ok"})

    # --- Lógica principal (POST) ---
    def do_POST(self):
        try:
            # 1) Leer JSON
            content_length = int(self.headers.get('Content-Length', 0))
            raw = self.rfile.read(content_length).decode('utf-8') if content_length else '{}'
            data = json.loads(raw)

            cantidad_pallets = int(data.get("cantidad_pallets", 0))
            meses_operacion = int(data.get("meses_operacion", 0))

            # Validación básica
            if meses_operacion < 1 or meses_operacion > 12:
                return self._json_response(400, {"error": "meses_operacion debe estar entre 1 y 12"})
            if cantidad_pallets < 0:
                return self._json_response(400, {"error": "cantidad_pallets debe ser >= 0"})

            # 2) Generar IN/OUT aleatorio (tu lógica)
            in_values = [0] * 12
            out_values = [0] * 12

            remaining_in = cantidad_pallets
            for i in range(meses_operacion):
                if i == meses_operacion - 1:
                    in_values[i] = remaining_in
                else:
                    denom = (meses_operacion - i)
                    max_val = remaining_in // denom if denom > 0 else remaining_in
                    val = random.randint(0, max_val)
                    in_values[i] = val
                    remaining_in -= val

            remaining_out = cantidad_pallets
            stock = 0
            for i in range(meses_operacion):
                stock += in_values[i]
                if i == meses_operacion - 1:
                    out_values[i] = remaining_out
                else:
                    denom = (meses_operacion - i)
                    max_out = min(stock, remaining_out // denom if denom > 0 else remaining_out)
                    val = random.randint(0, max_out)
                    out_values[i] = val
                    stock -= val
                    remaining_out -= val

            # 3) Rutas del Excel (leer del repo, escribir en /tmp)
            base_dir = os.path.dirname(__file__)
            src_xlsx = os.path.abspath(os.path.join(base_dir, '..', '..', 'simulacion.xlsx'))
            tmp_xlsx = '/tmp/simulacion.xlsx'

            # 4) Actualizar Excel
            wb = openpyxl.load_workbook(src_xlsx)
            ws = wb["cliente"]
            for i in range(12):
                ws[f"D{9+i}"] = in_values[i]
                ws[f"E{9+i}"] = out_values[i]
            wb.save(tmp_xlsx)

            # 5) Intentar recalcular con LibreOffice (si está disponible)
            recalculated_path = tmp_xlsx
            try:
                subprocess.run(
                    ["libreoffice", "--headless", "--calc", "--convert-to", "xlsx", "--outdir", "/tmp", tmp_xlsx],
                    check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=30
                )
                # Si se generó, úsalo
                if os.path.exists('/tmp/simulacion.xlsx'):
                    recalculated_path = '/tmp/simulacion.xlsx'
            except Exception:
                # Si falla, seguimos con data_only (sin recálculo de fórmulas)
                pass

            # 6) Leer datos finales (data_only=True)
            wb2 = openpyxl.load_workbook(recalculated_path, data_only=True)
            ws2 = wb2["cliente"]

            tarjetas = {
                "pallet_parking": ws2["P103"].value,
                "tradicional": ws2["P104"].value,
                "ahorro": ws2["P105"].value
            }
            tabla = []
            for i in range(meses_operacion):
                tabla.append({
                    "mes": i + 1,
                    "in": ws2[f"D{9+i}"].value,
                    "out": ws2[f"E{9+i}"].value,
                    "stock": ws2[f"G{9+i}"].value
                })
            costos_pp = [ws2.cell(row=103, column=4+i).value for i in range(12)]
            costos_trad = [ws2.cell(row=104, column=4+i).value for i in range(12)]

            resultado = {
                "tarjetas": tarjetas,
                "tabla": tabla,
                "costos": {
                    "pallet_parking": costos_pp,
                    "tradicional": costos_trad
                }
            }

            return self._json_response(200, resultado)

        except Exception as e:
            # Cualquier error inesperado
            return self._json_response(500, {"error": str(e)})
